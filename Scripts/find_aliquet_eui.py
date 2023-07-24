import os
import glob
from openpyxl import load_workbook
from datetime import datetime


def get_eui_dev_from_xlsx():

    PATH =  sorted(glob.glob(os.path.dirname(__file__)+'/*.xlsx'), key=os.path.getmtime)
    wb = load_workbook(PATH[-1].replace('  ', '_'))
    sheet = wb[wb.sheetnames[0]]
    DEV_LIST = []
    i=0
    while True:
        i+=1
        EUI_DEV = sheet[f'D{i}'].value
        if EUI_DEV == None:
            break
        try:
            DEV_LIST.append(EUI_DEV)
                
        except TypeError :
            break

    return DEV_LIST


def check_in_log_zeus():

    LIST_DEVICE = get_eui_dev_from_xlsx()
    string_find = '[INFO] Received from socket: 5501AB1D00'

    PATH =  sorted(glob.glob(os.path.dirname(__file__)+'/*.log'), key=os.path.getmtime)
    with open(PATH[-1], 'r', encoding='utf-8') as f:
        data = f.readlines()

    LIST_EUI = []

    for line in data:
        if line.find(string_find) != -1:
            LIST_EUI.append(line[line.find(string_find)+len(string_find):line.find(string_find)+len(string_find)+16])

    number = 1
    for i in set(LIST_DEVICE):
        if i in set(LIST_EUI):
            print('{:.<25} {:<10} {:<2}'.format(i, 'ОПРОШЕН', str(number)))
            # with open('opros_IVK.log', 'a') as f:
            #     print('{:.<25} {:<10} {:<2}'.format(i, 'ОПРОШЕН', str(number)), file=f)

        else:
            print('{:.<25} {:<10} {:<2}'.format(i, 'НЕ ОПРОШЕН', str(number)))
            with open('opros_IVK.log', 'a') as f:
                print('{:.<25} {:<10} {:<2}'.format(i, 'НЕ ОПРОШЕН', str(number)), file=f)

        number+=1
    # print('\n'.join(set(LIST_EUI)))

if __name__ == '__main__':

    check_in_log_zeus()